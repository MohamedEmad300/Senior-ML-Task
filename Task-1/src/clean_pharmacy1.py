"""Step B1: reshape Pharmacy 1's raw 2-header/2-data block export into one
row per item.

Verified against the real file (34,409 raw rows): every item is a perfectly
regular 5-row block relative to its "A" anchor --

    offset 0: A header row  (col0='الإجمالى' qty, col2 unit, col5 manufacturer, col10 name, col27 item code)
    offset 1: A data row    (same column positions, values)
    offset 2: blank
    offset 3: B header row  (col0='موقع الصنف' location, col4 expiry, col6 batch,
                              col10 purchase date, col12 sale price, col16 cost price,
                              col19 quantity, col24 supplier, col29 supplier code)
    offset 4: B data row    (same column positions, values)

followed by 1-3 blank spacer rows before the next block. There were 4,709
A-anchors and 4,709 B-anchors in the source file with zero mismatches, so we
scan for A-anchors and read fixed relative offsets, but still verify the B
header appears at the expected offset and log (not silently drop) any block
that doesn't match, rather than assuming this regularity holds forever.
"""
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import config

ANCHOR_A = "الإجمالى"
ANCHOR_B = "موقع الصنف"

COL_A_QTY = 0
COL_A_UNIT = 2
COL_A_MANUFACTURER = 5
COL_A_NAME = 10
COL_A_ITEM_CODE = 27

COL_B_LOCATION = 0
COL_B_EXPIRY = 4
COL_B_BATCH = 6
COL_B_PURCHASE_DATE = 10
COL_B_SALE_PRICE = 12
COL_B_COST_PRICE = 16
COL_B_QUANTITY = 19
COL_B_SUPPLIER = 24
COL_B_SUPPLIER_CODE = 29


def normalize_name(name: str) -> str:
    if not isinstance(name, str):
        return ""
    n = name.upper()
    n = re.sub(r"[^\w\s]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def _safe(row, idx):
    return row[idx] if idx < len(row) else None


def reshape_pharmacy1(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    a_rows = [i for i, r in enumerate(rows) if _safe(r, 0) == ANCHOR_A]

    records = []
    dropped = []
    for ai in a_rows:
        try:
            a_data = rows[ai + 1]
            b_header = rows[ai + 3]
            b_data = rows[ai + 4]
        except IndexError:
            dropped.append((ai, "block runs past end of sheet"))
            continue

        if _safe(b_header, 0) != ANCHOR_B:
            dropped.append((ai, f"expected B anchor at offset 3, got {_safe(b_header, 0)!r}"))
            continue

        name = _safe(a_data, COL_A_NAME)
        location_path = _safe(b_data, COL_B_LOCATION)
        category_hint = None
        if isinstance(location_path, str) and ">" in location_path:
            parts = [p.strip() for p in location_path.split(">")]
            if len(parts) >= 2:
                category_hint = parts[1]

        records.append({
            "Item_Name": name,
            "Name_Normalized": normalize_name(name),
            "Quantity": _safe(a_data, COL_A_QTY),
            "Unit": _safe(a_data, COL_A_UNIT),
            "Manufacturer": _safe(a_data, COL_A_MANUFACTURER),
            "Item_Code": _safe(a_data, COL_A_ITEM_CODE),
            "Location_Path": location_path,
            "Category_Hint": category_hint,
            "Expiry_Date": _safe(b_data, COL_B_EXPIRY),
            "Batch": _safe(b_data, COL_B_BATCH),
            "Purchase_Date": _safe(b_data, COL_B_PURCHASE_DATE),
            "Sale_Price": _safe(b_data, COL_B_SALE_PRICE),
            "Cost_Price": _safe(b_data, COL_B_COST_PRICE),
            "Supplier": _safe(b_data, COL_B_SUPPLIER),
            "Supplier_Code": _safe(b_data, COL_B_SUPPLIER_CODE),
        })

    if dropped:
        print(f"WARNING: {len(dropped)} block(s) dropped during Pharmacy 1 reshape:")
        for row_idx, reason in dropped[:20]:
            print(f"  row {row_idx + 1}: {reason}")
        if len(dropped) > 20:
            print(f"  ... and {len(dropped) - 20} more")

    df = pd.DataFrame.from_records(records)
    return df, dropped


def main():
    print(f"Reading {config.INPUT_PATH} ...")
    wb = openpyxl.load_workbook(config.INPUT_PATH, read_only=True, data_only=True)
    ws = wb[config.SHEET_PHARMACY1]
    print(f"Raw rows: {ws.max_row}")

    df, dropped = reshape_pharmacy1(ws)
    print(f"Reshaped into {len(df)} items ({len(dropped)} blocks dropped).")

    n_missing_name = df["Item_Name"].isna().sum()
    if n_missing_name:
        print(f"WARNING: {n_missing_name} items have no name.")

    out_path = config.CACHE_DIR / "pharmacy1_cleaned.pkl"
    df.to_pickle(out_path)
    print(f"Saved cleaned table to {out_path}")
    return df


if __name__ == "__main__":
    main()
